import json
from copy import copy

from django.db.models import Q
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import redirect, render
from django.template.loader import get_template
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import ListView, CreateView, DeleteView, UpdateView, FormView, DetailView
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment

from src.admin_panel.forms.receipts import ReceiptFilterForm, ReceiptForm, ReceiptServiceFormset, ReceiptExcelDocForm
from src.admin_panel.models import Receipt, Paybox, PersonalAccount, Indication, ReceiptService, Measure, Service, \
    ReceiptExcelDoc

from ..tasks import send_receipt


def balances(context):
    total_plus = sum(Paybox.objects.filter(debit_credit='plus', is_complete=True).values_list('total', flat=True))
    total_minus = sum(Paybox.objects.filter(debit_credit='minus', is_complete=True).values_list('total', flat=True))
    context['paybox_balance'] = total_plus - total_minus

    context['personal_accounts_debts'] = sum(
        PersonalAccount.objects.filter(balance__lt=0).values_list('balance', flat=True))

    personal_accounts_plus = sum(
        Paybox.objects.filter(debit_credit='plus', is_complete=True, personal_account__isnull=False).values_list(
            'total', flat=True))
    personal_accounts_minus = sum(
        Receipt.objects.filter(is_complete=True, flat__personal_account__isnull=False).values_list(
            'total_price', flat=True))
    context['personal_accounts_balance'] = personal_accounts_plus - personal_accounts_minus


class ReceiptList(ListView):
    template_name = 'receipt/receipts.html'
    context_object_name = 'receipts'
    queryset = Receipt.objects.all()
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ReceiptFilterForm()
        balances(context)
        return context


class CreateReceipt(CreateView):
    model = Receipt
    template_name = 'receipt/receipt_form.html'
    form_class = ReceiptForm
    success_url = reverse_lazy('receipts')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['indications'] = Indication.objects.order_by('date_published').all()
        context['service_formset'] = ReceiptServiceFormset(queryset=ReceiptService.objects.none(), prefix='service')
        context['measures'] = Measure.objects.all()
        context['services'] = Service.objects.all()
        balances(context)
        return context

    def post(self, request, *args, **kwargs):
        service_formset = ReceiptServiceFormset(request.POST, prefix='service')
        receipt_form = ReceiptForm(request.POST)
        if receipt_form.is_valid() and service_formset.is_valid():
            obj = receipt_form.save()
            instances = service_formset.save(commit=False)
            for instance in instances:
                instance.receipt_id = obj.id
                instance.save()
            if hasattr(obj.flat, 'personal_account'):
                if obj.flat.personal_account is not None and obj.flat.personal_account != '' and obj.is_complete is True:
                    personal_account = PersonalAccount.objects.get(pk=obj.flat.personal_account.id)
                    personal_account.balance = personal_account.balance - obj.total_price
                    personal_account.save()
                return redirect('receipts')
        else:
            context = {
                "indications": Indication.objects.order_by('date_published').all(),
                "service_formset": service_formset,
                "form": receipt_form,
                'measures': Measure.objects.all(),
                'services': Service.objects.all(),
            }
            return render(request, self.template_name, context)


class GetIndicationsSortedList(ListView):
    def get(self, request, flat_id):
        indications = Indication.objects.order_by('date_published').filter(flat_id=flat_id)
        context = {
            'indications': indications,
        }
        return render(request, 'receipt/indications_table.html', context)


class ReceiptsFilteredList(ListView):
    template_name = 'receipt/receipts.html'
    context_object_name = 'receipts'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ReceiptFilterForm(initial=self.request.GET)
        balances(context)
        return context

    def get_queryset(self):
        receipts = Receipt.objects.all()
        form_filter = ReceiptFilterForm(self.request.GET)
        qs = []
        if form_filter.is_valid():
            if form_filter.cleaned_data['number']:
                qs.append(Q(number__icontains=form_filter.cleaned_data['number']))
            if form_filter.cleaned_data['status']:
                qs.append(Q(status=form_filter.cleaned_data['status']))
            if form_filter.cleaned_data['daterange']:
                date_start, date_end = str(form_filter.cleaned_data['daterange']).split(' - ')
                date_start = date_start.split('/')
                date_end = date_end.split('/')
                date_start.reverse()
                date_end.reverse()
                date_start = "-".join(date_start)
                date_end = "-".join(date_end)
                qs.append(Q(
                    Q(date_published__gte=date_start) &
                    Q(date_published__lte=date_end)
                ))
            if form_filter.cleaned_data['month']:
                date_list = form_filter.cleaned_data['month'].split('-')
                date_list.reverse()
                result = "-".join(date_list)
                qs.append(Q(date_published__month=date_list[1]))
            if form_filter.cleaned_data['flat']:
                qs.append(Q(flat__number__icontains=form_filter.cleaned_data['flat']))
            if form_filter.cleaned_data['flat_owner']:
                full_name = str(form_filter.cleaned_data['flat_owner']).split(' ')
                qs.append(Q(
                    Q(flat__flat_owner__user__last_name__icontains=full_name[0]) &
                    Q(flat__flat_owner__patronymic__icontains=full_name[2]) &
                    Q(flat__flat_owner__user__first_name__icontains=full_name[1])
                ))

            if form_filter.cleaned_data['complete']:
                if form_filter.cleaned_data['complete'] == 'complete':
                    qs.append(Q(is_complete=True))
                if form_filter.cleaned_data['complete'] == 'no complete':
                    qs.append(Q(is_complete=False))
            q = Q()
            for item in qs:
                q = q & item
            receipts = Receipt.objects.filter(q)
        return receipts


class SendReceiptEmail(View):
    def get(self, request, receipt_id, *args, **kwargs):
        receipt = Receipt.objects.get(pk=receipt_id)
        services = ReceiptService.objects.filter(receipt_id=receipt_id)
        context = {
            'services': services,
            'receipt': receipt,
        }
        html_template = get_template('cabinet/invoice.html')
        html = html_template.render(context)
        base_url = request.build_absolute_uri()

        send_receipt.delay(html_to_pdf=html, base_url=base_url,
                           to=receipt.flat.flat_owner.user.email)
        url = f"/admin/receipt/detail/{receipt_id}"
        return HttpResponseRedirect(url)


class UpdateReceipt(UpdateView):
    model = Receipt
    template_name = 'receipt/receipt_update.html'
    form_class = ReceiptForm
    success_url = reverse_lazy('receipts')

    def get(self, request, *args, **kwargs):
        receipt = Receipt.objects.get(id=self.kwargs['pk'])
        context = {
            'indications': Indication.objects.order_by('date_published').filter(flat_id=receipt.flat.id),
            'form': ReceiptForm(instance=receipt,
                                initial={'house': receipt.flat.house, 'section': receipt.flat.section}),
            'service_formset': ReceiptServiceFormset(prefix='service', queryset=ReceiptService.objects.filter(
                receipt_id=self.kwargs['pk'])),
            'services': Service.objects.all(),
            'measures': Measure.objects.all(),
        }
        return render(request, 'receipt/receipt_update.html', context)

    def post(self, request, pk, *args, **kwargs):
        service_formset = ReceiptServiceFormset(request.POST, prefix='service')
        receipt_form = ReceiptForm(request.POST, instance=Receipt.objects.get(id=pk))

        if receipt_form.is_valid() and service_formset.is_valid():
            obj = receipt_form.save()
            instances = service_formset.save()
            for instance in instances:
                instance.receipt_id = obj.id
                instance.save()

            if hasattr(obj.flat, 'personal_account'):
                if obj.flat.personal_account is not None \
                        and obj.flat.personal_account != '' \
                        and obj.is_complete is True:
                    personal_account = PersonalAccount.objects.get(pk=obj.flat.personal_account.id)
                    plus_total = sum(
                        Paybox.objects.filter(personal_account=personal_account,
                                              is_complete=True).values_list('total', flat=True))
                    minus_total = sum(
                        Receipt.objects.filter(flat__personal_account=personal_account,
                                               is_complete=True).values_list('total_price', flat=True))

                    personal_account.balance = plus_total - minus_total
                    personal_account.save()
            return redirect('receipts')
        else:
            context = {
                "indications": Indication.objects.order_by('date_published').all(),
                "service_formset": service_formset,
                "form": receipt_form,
                'measures': Measure.objects.all(),
                'services': Service.objects.all(),
            }
            return render(request, 'receipt/receipt_update.html', context)


class ReceiptPrint(View):
    def get(self, request, pk, *args, **kwargs):
        receipt = Receipt.objects.get(pk=pk)
        rows = ReceiptExcelDoc.objects.all().order_by('id')
        data = {
            'rows': rows,
            'receipt': receipt
        }
        return render(request, 'receipt/receipt_printing.html', context=data)


def copy_row(ws, source_row_num, target_row_num):
    # Get the source row
    source_row = ws[source_row_num]

    # Copy the source row to a new row number
    for cell in source_row:
        if cell.coordinate in ws.merged_cells:
            for merged_cell in list(ws.merged_cells.ranges):
                if cell.coordinate in merged_cell:
                    merged_range = ws.cell(row=target_row_num,
                                           column=merged_cell.min_col).coordinate + ":" + \
                                   ws.cell(row=target_row_num,
                                           column=merged_cell.max_col).coordinate
                    ws.merge_cells(merged_range)
                    ws.cell(row=target_row_num, column=merged_cell.min_col, value=cell.value)
                    ws.cell(row=target_row_num, column=merged_cell.min_col).alignment = Alignment(horizontal='left',
                                                                                                  vertical='center')  # выровнять по центр
        else:
            ws.cell(row=target_row_num, column=cell.col_idx, value=cell.value)

    target_row = ws[target_row_num]  # Get the destination row

    # Copy styles from source row to destination row
    for source_cell, dest_cell in zip(source_row, target_row):
        dest_cell.font = copy(source_cell.font)  # Copy font style
        dest_cell.fill = copy(source_cell.fill)  # Copy fill style
        dest_cell.border = copy(source_cell.border)  # Copy border style


class ReceiptDownloadExcel(View):
    def post(self, request, excel_id, receipt_id, *args, **kwargs):
        receipt = Receipt.objects.get(pk=receipt_id)
        services = ReceiptService.objects.filter(receipt_id=receipt_id)
        services_count = services.count()
        doc = ReceiptExcelDoc.objects.get(pk=excel_id)

        wb = load_workbook(doc.file)
        ws = wb.active  # это лист в excel
        if services_count > 0:
            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=50):
                for cell in row:
                    if cell.value == 'total':
                        final_row = ws[cell.row]
                        ws.delete_rows(cell.row)

            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=50):
                for cell in row:
                    match cell.value:
                        case 'service':
                            for i, obj in enumerate(services):
                                ws.cell(row=cell.row + i, column=cell.column, value=str(obj.service.title))
                                copy_row(ws, cell.row, cell.row + i)
                                ws.cell(row=cell.row + i, column=cell.column, value=str(obj.service.title))
                            else:
                                for cell in final_row:
                                    if cell.coordinate in ws.merged_cells:
                                        for merged_cell in list(ws.merged_cells.ranges):
                                            if cell.coordinate in merged_cell:
                                                merged_range = ws.cell(row=cell.row + i,
                                                                       column=merged_cell.min_col).coordinate + ":" + \
                                                               ws.cell(row=cell.row + i,
                                                                       column=merged_cell.max_col).coordinate
                                                ws.merge_cells(merged_range)
                                                ws.cell(row=cell.row + i, column=merged_cell.min_col, value=cell.value)
                                                ws.cell(row=cell.row + i,
                                                        column=merged_cell.min_col).alignment = Alignment(
                                                    horizontal='right', vertical='center')  # выровнять по центр
                                    else:
                                        ws.cell(row=cell.row + i, column=cell.col_idx, value=cell.value)
                                for source_cell, dest_cell in zip(final_row, ws[cell.row + i]):
                                    dest_cell.font = copy(source_cell.font)  # Copy font style
                                    dest_cell.fill = copy(source_cell.fill)  # Copy fill style
                                    dest_cell.border = copy(source_cell.border)  # Copy border style
                        case 'tariff':
                            for i, obj in enumerate(services):
                                ws.cell(row=cell.row + i, column=cell.column, value=str(obj.receipt.tariff.title))
                        case 'measure':
                            for i, obj in enumerate(services):
                                ws.cell(row=cell.row + i, column=cell.column, value=str(obj.measure.title))
                        case 'totalServicePrice':
                            for i, obj in enumerate(services):
                                ws.cell(row=cell.row + i, column=cell.column,
                                        value=str(obj.unit_price * obj.consumption))
        else:
            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=50):
                for cell in row:
                    match cell.value:
                        case 'service':
                            ws.delete_rows(cell.row)

        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=50):
            for cell in row:
                match cell.value:
                    case 'personal_accountNumber':
                        if hasattr(receipt.flat, 'personal_account'):
                            if receipt.flat.personal_account is None or receipt.flat.personal_account == "":
                                cell.value = ''
                            else:
                                cell.value = receipt.flat.personal_account.number
                        else:
                            cell.value = ''
                    case 'personalManager':
                        if receipt.flat.flat_owner is None:
                            cell.value = ''
                        else:
                            cell.value = str(receipt.flat.flat_owner)
                            cell.alignment = Alignment(horizontal='center', vertical='center')  # выровнять по центру
                    case 'receiptNumber':
                        cell.value = receipt.number
                    case 'receiptStartDate':
                        cell.value = str(receipt.start_date)
                    case 'totalPrice':
                        cell.value = receipt.total_price
                    case 'flatOwner':
                        if receipt.flat.flat_owner is None:
                            cell.value = ''
                        else:
                            cell.value = str(receipt.flat.flat_owner)
                    case 'personalAccountBalance':
                        if hasattr(receipt.flat, 'personal_account'):
                            if receipt.flat.personal_account is None or receipt.flat.personal_account == "":
                                cell.value = ''
                            else:
                                cell.value = receipt.flat.personal_account.balance
                        else:
                            cell.value = ''
                    case 'receiptDatePublished':
                        cell.value = receipt.date_published
                        cell.alignment = Alignment(horizontal='center', vertical='center')  # выровнять по центр
                    case 'receiptMonthPublished':
                        month_dict = {
                            'January': 'Январь',
                            'February': 'Февраль',
                            'March': 'Март',
                            'April': 'Апрель',
                            'May': 'Май',
                            'June': 'Июнь',
                            'July': 'Июль',
                            'August': 'Август',
                            'September': 'Сентябрь',
                            'October': 'Октябрь',
                            'November': 'Ноябрь',
                            'December': 'Декабрь'
                        }
                        cell.value = month_dict[str(receipt.date_published.strftime('%B'))]
                        cell.alignment = Alignment(horizontal='center', vertical='center')  # выровнять по центр
                    case 'total':
                        cell.value = receipt.total_price
        ws.title = "Квитанция"  # это название листа в excel
        wb.save('media/receipt/result/report.xlsx')

        data = {
        }
        return HttpResponse(json.dumps(data), content_type='application/json')


class ReceiptPrintingSettings(View):
    def get(self, request, *args, **kwargs):
        rows = ReceiptExcelDoc.objects.all().order_by('id')
        form = ReceiptExcelDocForm()
        data = {
            'rows': rows,
            'form': form,
        }
        return render(request, 'receipt/receipt_printing_settings.html', context=data)

    def post(self, request, *args, **kwargs):
        form = ReceiptExcelDocForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
        return redirect('receipt_print_settings')


class ReceiptPrintingSettingsDefault(View):
    def get(self, request, pk, *args, **kwargs):
        for row in ReceiptExcelDoc.objects.all():
            row.by_default = False
            row.save()
        row = ReceiptExcelDoc.objects.get(pk=pk)
        row.by_default = True
        row.save()
        return redirect('receipt_print_settings')


class ReceiptPrintingSettingsDelete(DeleteView):
    success_url = reverse_lazy('receipt_print_settings')
    queryset = ReceiptExcelDoc.objects.all()


class CopyReceipt(FormView):
    def get(self, request, pk, *args, **kwargs):
        copy = Receipt.objects.get(pk=pk)
        receipt_form = ReceiptForm(instance=copy, initial={'house': copy.flat.house, 'section': copy.flat.section})

        service_formset = ReceiptServiceFormset(queryset=ReceiptService.objects.filter(receipt_id=pk),
                                                prefix='service')
        data = {
            'form': receipt_form,
            'service_formset': service_formset,
            "indications": Indication.objects.order_by('date_published').filter(flat_id=copy.flat.id),
            'measures': Measure.objects.all(),
            'services': Service.objects.all(),
        }
        return render(request, 'receipt/receipt_form.html', context=data)

    def post(self, request, *args, **kwargs):
        service_formset = ReceiptServiceFormset(request.POST, prefix='service')
        receipt_form = ReceiptForm(request.POST)
        if receipt_form.is_valid() and service_formset.is_valid():
            obj = receipt_form.save()
            instances = service_formset.save(commit=False)
            for instance in instances:
                instance.receipt_id = obj.id
                instance.save()
            return redirect('receipts')
        else:
            data = {
                "indications": Indication.objects.order_by('date_published').all(),
                "service_formset": service_formset,
                "form": receipt_form,
                'measures': Measure.objects.all(),
                'services': Service.objects.all(),
            }
            return render(request, 'receipt/receipt_form.html', context=data)


class ReceiptDetail(DetailView):
    model = Receipt
    template_name = 'receipt/receipt_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        receipt = Receipt.objects.get(pk=self.kwargs['pk'])
        receipt_services = ReceiptService.objects.filter(receipt_id=self.kwargs['pk'])
        context['receipt'] = receipt
        context['receipt_services'] = receipt_services
        return context


class DeleteReceipt(DeleteView):
    success_url = reverse_lazy('receipts')
    queryset = Receipt.objects.all()

    def post(self, request, *args, **kwargs):
        receipt = Receipt.objects.get(pk=self.kwargs['pk'])
        if hasattr(receipt.flat, 'personal_account'):
            if receipt.flat.personal_account is not None \
                    and receipt.flat.personal_account != '' \
                    and receipt.is_complete is True:
                personal_account = PersonalAccount.objects.get(pk=receipt.flat.personal_account.id)
                receipt.delete()
                plus_total = sum(
                    Paybox.objects.filter(personal_account=personal_account,
                                          is_complete=True).values_list('total', flat=True))
                minus_total = sum(
                    Receipt.objects.filter(flat__personal_account=personal_account,
                                           is_complete=True).values_list('total_price', flat=True))

                personal_account.balance = plus_total - minus_total
                personal_account.save()
        else:
            receipt.delete()
        return redirect('receipts')
