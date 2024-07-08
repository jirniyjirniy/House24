from django.db.models import Q
from django.shortcuts import render, redirect
from django.utils import timezone
from django.views.generic import ListView, FormView, DetailView
from openpyxl.styles import Font
from openpyxl.workbook import Workbook

from src.admin_panel.forms.paybox import PayboxFilterForm, PayboxForm
from src.admin_panel.models import Paybox, Personal, Article, PersonalAccount, Receipt


def balances(context):
    total_plus = sum(Paybox.objects.filter(debit_credit='plus', is_complete=True).values_list('total', flat=True))
    total_minus = sum(Paybox.objects.filter(debit_credit='minus', is_complete=True).values_list('total', flat=True))
    context['paybox_balance'] = total_plus - total_minus

    context['personal_accounts_debts'] = sum(
        PersonalAccount.objects.filter(balance__lt=0).values_list('balance', flat=True))

    personal_accounts_plus = sum(
        Paybox.objects.filter(debit_credit='plus', is_complete=True, personal_account__isnull=False).values_list(
            'total', flat=True))
    personal_accounts_minus = sum(
        Receipt.objects.filter(is_complete=True, flat__personal_account__isnull=False).values_list(
            'total_price', flat=True))
    context['personal_accounts_balance'] = personal_accounts_plus - personal_accounts_minus


class PayboxList(ListView):
    template_name = 'paybox/paybox.html'
    context_object_name = 'paybox'
    paginate_by = 20

    def get_queryset(self):
        paybox = Paybox.objects.all()
        wb = Workbook()
        ws = wb.active  # это лист в excel
        ws.append(
            [
                '#',
                'Дата',
                'Приход/расход',
                'Статус',
                'Статья',
                'Квитанция',
                'Услуга',
                'Сумма',
                'Валюта',
                'Владелец квартиры',
                'Лицевой счет',
            ]
        )
        bold_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold_font

        for obj in paybox:
            if obj.flat_owner is not None:
                flat_owner = f'{obj.flat_owner}'
            else:
                flat_owner = f''

            if obj.personal_account is not None:
                personal_account = f'{obj.personal_account}'
            else:
                personal_account = f''
            ws.append([
                f'{obj.number}',
                f'{obj.date_published}',
                f'{obj.get_debit_credit_display()}',
                f'{"Проведен" if obj.is_complete else "Не проведен"}',
                f'{obj.article.title}',
                f'',
                f'',
                f'{obj.total}',
                f'UAH',
                f'{flat_owner}',
                f'{personal_account}',
            ])

        ws.title = "Выписка"  # это название листа в excel
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 40
        ws.column_dimensions['K'].width = 20

        wb.save('media/paybox/all_info.xlsx')

        # Get full path to workbook
        wb.close()

        return paybox

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        total_plus = sum(Paybox.objects.filter(debit_credit='plus', is_complete=True).values_list('total', flat=True))
        total_minus = sum(Paybox.objects.filter(debit_credit='minus', is_complete=True).values_list('total', flat=True))
        context['filter_form'] = PayboxFilterForm()
        context['total_plus'] = total_plus
        context['total_minus'] = total_minus
        balances(context)
        return context


class PayboxFilteredList(ListView):
    template_name = 'paybox/paybox.html'
    context_object_name = 'paybox'
    paginate_by = 20

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = PayboxFilterForm(initial=self.request.GET)
        return context

    def get_queryset(self):
        paybox = Paybox.objects.all()
        form_filter = PayboxFilterForm(self.request.GET)
        qs = []
        if form_filter.is_valid():
            if form_filter.cleaned_data['number']:
                qs.append(Q(number__icontains=form_filter.cleaned_data['number']))
            if form_filter.cleaned_data['personal_account']:
                qs.append(Q(personal_account__number__icontains=form_filter.cleaned_data['personal_account']))
            if form_filter.cleaned_data['daterange']:
                date_start, date_end = str(form_filter.cleaned_data['daterange']).split(' - ')
                date_start = date_start.split('/')
                date_end = date_end.split('/')
                date_start.reverse()
                date_end.reverse()
                date_start = "-".join(date_start)
                date_end = "-".join(date_end)
                qs.append(Q(
                    Q(date_published__gte=date_start) &
                    Q(date_published__lte=date_end)
                ))

            if form_filter.cleaned_data['article']:
                qs.append(Q(article_id=form_filter.cleaned_data['article'].id))
            if form_filter.cleaned_data['flat_owner']:
                full_name = str(form_filter.cleaned_data['flat_owner']).split(' ')
                qs.append(Q(
                    Q(flat_owner__user__last_name__icontains=full_name[0]) &
                    Q(flat_owner__patronymic__icontains=full_name[2]) &
                    Q(flat_owner__user__first_name__icontains=full_name[1])
                ))

            if form_filter.cleaned_data['status']:
                if form_filter.cleaned_data['status'] == 'complete':
                    qs.append(Q(is_complete=True))
                if form_filter.cleaned_data['status'] == 'no complete':
                    qs.append(Q(is_complete=False))
            if form_filter.cleaned_data['debit_credit']:
                qs.append(Q(debit_credit=form_filter.cleaned_data['debit_credit']))
            q = Q()
            for item in qs:
                q = q & item
            paybox = Paybox.objects.filter(q)
        return paybox


class CreatePaybox(FormView):
    def get(self, request, income, *args, **kwargs):
        form = PayboxForm()
        form.fields['date_published'].initial = timezone.now().date()
        form.fields['user'].initial = Personal.objects.get(user_id=self.request.user.id)
        form.fields['is_complete'].initial = True

        if income == 'plus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="plus")
        elif income == 'minus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="minus")
        data = {
            'income': income,
            'form': form,
        }
        return render(request, 'paybox/paybox_form.html', context=data)

    def post(self, request, income, *args, **kwargs):
        form = PayboxForm(request.POST)

        if form.is_valid():
            instance = form.save()
            if income == 'plus':
                if instance.personal_account is not None and instance.is_complete is True:
                    personal_account = PersonalAccount.objects.get(pk=instance.personal_account_id)
                    personal_account.balance = personal_account.balance + instance.total
                    personal_account.save()
                instance.debit_credit = 'plus'
            elif income == 'minus':
                instance.debit_credit = 'minus'
            instance.save()
            return redirect('paybox')
        else:

            data = {
                'income': income,
                'form': form,
            }
            return render(request, 'paybox/paybox_form.html', context=data)


class CopyPaybox(CreatePaybox):
    def get(self, request, pk, *args, **kwargs):
        copy = Paybox.objects.get(pk=pk)
        form = PayboxForm(instance=copy)

        if copy.debit_credit == 'plus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="plus")
        elif copy.debit_credit == 'minus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="minus")
        income = copy.debit_credit
        data = {
            'income': income,
            'form': form,
        }
        return render(request, 'paybox/paybox_form.html', context=data)


from urllib.parse import urlencode, quote


class PayboxDetail(DetailView):
    model = Paybox
    template_name = 'paybox/paybox_detail.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        paybox = Paybox.objects.get(pk=self.kwargs['pk'])
        context['paybox'] = paybox

        wb = Workbook()
        ws = wb.active
        ws.append(['Платёж', f'#{paybox.number}'])
        ws.append(['Дата', f'{paybox.date_published}'])
        if paybox.flat_owner:
            ws.append(['Владелец квартиры',
                       f'{paybox.flat_owner.user.last_name} {paybox.flat_owner.user.first_name} {paybox.flat_owner.patronymic}'])
        else:
            ws.append(['Владелец квартиры', f'не указан'])

        if paybox.personal_account:
            ws.append(['Лицевой счёт', paybox.personal_account.number])
        else:
            ws.append(['Лицевой счёт', f'не указан'])

        ws.append(['Приход/Расход', paybox.get_debit_credit_display()])
        ws.append(['Проведён', 'Проведён' if paybox.is_complete else 'Не проведён'])
        if paybox.article:
            ws.append(['Статья', paybox.article.title])
        else:
            ws.append(['Статья', f'не указана'])

        ws.append(['Квитанция', ''])
        ws.append(['Услуга', ''])
        ws.append(['Сумма', f'{paybox.total}'])
        ws.append(['Валюта', 'UAH'])
        ws.append(['Комментарий', paybox.comment])
        ws.append(['Приход/Расход', paybox.get_debit_credit_display()])
        ws.append(['Проведён', 'Проведён' if paybox.is_complete else 'Не проведён'])
        if paybox.user:
            ws.append(['Менеджер', f'{paybox.user.user.last_name} {paybox.user.user.first_name}'])
        else:
            ws.append(['Менеджер', f'не указан'])
        ws.title = "Выписка"  # это название листа в excel
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

        wb.save('media/paybox/info.xlsx')

        # Get full path to workbook
        wb.close()

        return context


class UpdatePaybox(FormView):
    def get(self, request, pk, *args, **kwargs):
        paybox = Paybox.objects.get(pk=pk)
        form = PayboxForm(instance=paybox)
        if paybox.debit_credit == 'plus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="plus")
        elif paybox.debit_credit == 'minus':
            form.fields['article'].queryset = Article.objects.filter(debit_credit="minus")
        data = {
            'form': form,
        }
        return render(request, 'paybox/paybox_update.html', context=data)

    def post(self, request, pk, *args, **kwargs):
        paybox = Paybox.objects.get(pk=pk)
        form = PayboxForm(request.POST, instance=paybox)
        if form.is_valid():
            instance = form.save()
            if instance.debit_credit == 'plus':

                if instance.personal_account is not None and instance.is_complete is True:
                    personal_account = PersonalAccount.objects.get(pk=instance.personal_account_id)
                    plus_total = sum(
                        Paybox.objects.filter(personal_account=personal_account,
                                              is_complete=True).values_list('total', flat=True))
                    minus_total = sum(
                        Receipt.objects.filter(flat__personal_account=personal_account,
                                               is_complete=True).values_list('total_price', flat=True))

                    personal_account.balance = plus_total - minus_total
                    personal_account.save()
            return redirect('paybox')
        else:
            data = {
                'form': form,
            }
            return render(request, 'paybox/paybox_update.html', context=data)


class DeletePaybox(FormView):
    def post(self, request, pk, *args, **kwargs):
        paybox = Paybox.objects.get(pk=pk)
        if paybox.debit_credit == 'plus':
            if paybox.personal_account is not None and paybox.is_complete is True:
                personal_account = PersonalAccount.objects.get(pk=paybox.personal_account_id)
                paybox.delete()
                plus_total = sum(
                    Paybox.objects.filter(personal_account=personal_account,
                                          is_complete=True).values_list('total', flat=True))
                minus_total = sum(
                    Receipt.objects.filter(flat__personal_account=personal_account,
                                           is_complete=True).values_list('total_price', flat=True))

                personal_account.balance = plus_total - minus_total
                personal_account.save()
            else:
                paybox.delete()

        else:
            paybox.delete()
        return redirect('paybox')
